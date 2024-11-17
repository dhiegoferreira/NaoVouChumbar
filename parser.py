import openpyxl
from datetime import datetime, timedelta

def parse_cell(cell_value):
    events = []
    for line in cell_value.split('\n'):
        if line.strip():
            time, *description = line.split(' ', 1)
            events.append({
                'time': time,
                'description': description[0] if description else ''
            })
    return events

def parse_schedule(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    schedule = []

    for row in range(2, sheet.max_row + 1):
        week = sheet.cell(row=row, column=1).value
        if week and isinstance(week, int):
            for day in range(2, 8):  # Monday to Saturday
                date_cell = sheet.cell(row=row, column=day)
                if date_cell.value:
                    date_str = date_cell.value.split('\n')[1]
                    date = datetime.strptime(f"{date_str}/2024", "%d/%b/%Y")
                    
                    for year in range(3):
                        cell = sheet.cell(row=row+year, column=day)
                        if cell.value:
                            events = parse_cell(cell.value)
                            for event in events:
                                schedule.append({
                                    'date': date,
                                    'year': year + 1,
                                    'time': event['time'],
                                    'description': event['description']
                                })
    return schedule

def main():
    file_path = 'data.xlsx'
    schedule = parse_schedule(file_path)
    
    for event in schedule:
        print(f"Date: {event['date'].strftime('%Y-%m-%d')}, "
              f"Year: {event['year']}, "
              f"Time: {event['time']}, "
              f"Description: {event['description']}")

if __name__ == "__main__":
    main()